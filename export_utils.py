from __future__ import annotations
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_to_tuple

def dfs_to_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    """
    Exporta varias hojas a un solo .xlsx.
    'sheets' es un dict: {"NombreHoja": DataFrame, ...}
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=name)
    return output.getvalue()

def df_to_excel_at_cell_bytes(df: pd.DataFrame, start_cell: str = "C7", sheet_name: str = "Comparativa") -> bytes:
    """Escribe un DataFrame en una hoja nueva empezando EXACTAMENTE en start_cell (incluye encabezado)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    row0, col0 = coordinate_to_tuple(start_cell)  # (fila, columna)

    # Encabezados
    for j, h in enumerate(df.columns):
        ws.cell(row=row0, column=col0 + j, value=str(h))

    # Datos
    for i, (_, r) in enumerate(df.iterrows(), start=1):
        for j, v in enumerate(r.tolist()):
            ws.cell(row=row0 + i, column=col0 + j, value=None if pd.isna(v) else v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def dfs_to_excel_multi_at_cell_bytes(sheets: dict[str, pd.DataFrame], start_cell: str = "C7") -> bytes:
    """
    Crea un .xlsx con varias hojas. Cada DF se escribe desde start_cell (incluye encabezado).
    sheets = {"Nuevos": df_nuevos, "Mantenidos": df_mant, "Eliminados": df_elim}
    """
    wb = Workbook()
    first = True
    row0, col0 = coordinate_to_tuple(start_cell)

    for sheet_name, df in sheets.items():
        ws = wb.active if first else wb.create_sheet(title=sheet_name)
        if first:
            ws.title = sheet_name
            first = False

        # Encabezados
        for j, h in enumerate(df.columns):
            ws.cell(row=row0, column=col0 + j, value=str(h))

        # Datos
        for i, (_, r) in enumerate(df.iterrows(), start=1):
            for j, v in enumerate(r.tolist()):
                ws.cell(row=row0 + i, column=col0 + j, value=None if pd.isna(v) else v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
